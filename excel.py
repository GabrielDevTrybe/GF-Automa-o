import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def processar_arquivo_hdi(caminho_arquivo_hdi, caminho_arquivo_destino):
    """Processar o arquivo HDI 431 e atualizar o arquivo de destino conforme requisitos."""
    try:
        # Carregar o workbook e a planilha ativa
        wb_hdi = load_workbook(caminho_arquivo_hdi)
        ws_hdi = wb_hdi.active

        # Excluir colunas A, B e D
        ws_hdi.delete_cols(1)  # Exclui a coluna A
        ws_hdi.delete_cols(1)  # Exclui a coluna B (agora a nova coluna A após a primeira exclusão)
        ws_hdi.delete_cols(2)  # Exclui a coluna D (agora a nova coluna B após as duas exclusões anteriores)

        # Inserir nova coluna na posição B (agora a coluna 2)
        ws_hdi.insert_cols(2)
        ws_hdi['B3'] = 'Nova Coluna'  # Opcional: Adiciona um título à nova coluna

        # Salvar alterações temporárias no arquivo HDI
        wb_hdi.save(caminho_arquivo_hdi)

        # Recarregar o arquivo HDI atualizado para processamento com pandas
        df_hdi = pd.read_excel(caminho_arquivo_hdi)

        # Função para limpar e formatar o valor
        def limpar_valor(valor):
            if pd.isna(valor):
                return ""
            # Remove espaços e caracteres especiais
            valor = ''.join(e for e in str(valor) if e.isalnum())
            return valor.upper()

        # Função para aplicar a fórmula na nova coluna B
        def aplicar_formula(valor, valor_anterior=None):
            valor = limpar_valor(valor)
            if valor_anterior and ' ' in valor:
                return limpar_valor(valor_anterior) + "-CORRETORA"
            return valor + '-CORRETORA'

        # Aplicar a fórmula diretamente na coluna B
        df_hdi['Nova_Coluna_B'] = df_hdi.iloc[:, 0].shift(1).fillna('').apply(lambda x: aplicar_formula(x))
        df_hdi['Nova_Coluna_B'] = df_hdi.apply(lambda row: aplicar_formula(row[df_hdi.columns[0]], row['Nova_Coluna_B']), axis=1)

        # Atualizar a coluna B do DataFrame
        df_hdi[df_hdi.columns[1]] = df_hdi['Nova_Coluna_B']

        # Excluir a coluna auxiliar 'Nova_Coluna_B'
        df_hdi.drop(columns=['Nova_Coluna_B'], inplace=True)

        # Função para formatar valores monetários
        def formatar_valor(valor):
            try:
                # Remove qualquer caractere não numérico, exceto o ponto e a vírgula
                valor = str(valor).replace('R$', '').replace('.', '').replace(',', '.')
                # Converte para float e formata
                valor_formatado = f'R$ {float(valor):,.2f}'.replace('.', ',')
                return valor_formatado
            except ValueError:
                return valor

        # Aplicar a formatação na coluna que precisa de valores monetários
        df_hdi[df_hdi.columns[1]] = df_hdi[df_hdi.columns[1]].apply(formatar_valor)

        # Salvar alterações no arquivo HDI
        df_hdi.to_excel(caminho_arquivo_hdi, index=False)
        print(f"Arquivo {caminho_arquivo_hdi} atualizado com sucesso!")

        # Carregar o arquivo de destino
        wb_destino = load_workbook(caminho_arquivo_destino)
        ws_destino = wb_destino.active

        # Encontrar a última linha preenchida na planilha de destino
        ultima_linha = ws_destino.max_row

        # Recarregar o arquivo HDI atualizado para copiar as colunas B e C
        df_hdi = pd.read_excel(caminho_arquivo_hdi, header=None)
        colunas_bc = df_hdi.iloc[3:, [1, 2]]  # A partir da linha 4 (índice 3) e colunas B e C

        # Adicionar os dados das colunas B e C ao final do arquivo TESTE
        for r in dataframe_to_rows(colunas_bc, index=False, header=False):
            ws_destino.append(r)

        # Salvar o arquivo de destino atualizado
        wb_destino.save(caminho_arquivo_destino)
        print(f"Arquivo {caminho_arquivo_destino} atualizado com sucesso!")

    except Exception as e:
        print(f"Erro ao processar os arquivos: {e}")

# Definição dos caminhos dos arquivos
arquivo_hdi_431 = 'C:\\Users\\User\\Desktop\\HDI\\HDI 431.xlsx'
arquivo_hdi_destino = 'C:\\Users\\User\\Desktop\\HDI\\TESTE.xlsx'

# Executar o processo
processar_arquivo_hdi(arquivo_hdi_431, arquivo_hdi_destino)
